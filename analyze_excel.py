
import pandas as pd
import openpyxl
import os

file_path = r"d:\src\simap-frontend\public\Salinan Presensi Kehadiran Tutor Ekstrakurikuler SMA (Jawaban).xlsx"

print(f"Analyzing file: {file_path}")

try:
    # 1. Load Workbook using openpyxl to get sheet names and basic structure with formatting in mind
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Sheet Names: {wb.sheetnames}")
    
    sheet = wb.active
    print(f"Active Sheet: {sheet.title}")
    
    # 2. Inspect first 15 rows to find headers and content
    print("\n--- First 15 Rows Content ---")
    for i, row in enumerate(sheet.iter_rows(max_row=15, values_only=True), 1):
        # Filter out None values for cleaner output
        row_content = [str(cell) if cell is not None else "" for cell in row]
        # Only print rows that are not completely empty
        if any(row_content):
            print(f"Row {i}: {row_content}")

    # 3. Analyze Column Headers (guessing based on content)
    # Usually headers are strictly string and contained in a single row
    
    print("\n--- Column Analysis ---")
    # Let's try to detect the header row. It often contains 'Nama', 'Tanggal', etc.
    header_keywords = ['nama', 'tutor', 'tanggal', 'hari', 'jam', 'waktu', 'materi', 'ttd', 'paraf']
    
    header_row_index = None
    for i, row in enumerate(sheet.iter_rows(max_row=15, values_only=True), 1):
        row_str = " ".join([str(c).lower() for c in row if c]).lower()
        if any(k in row_str for k in header_keywords):
            header_row_index = i
            print(f"Potential Header found at Row {i}: {row}")
            # Identify columns
            for idx, col_name in enumerate(row):
                if col_name:
                    print(f"  Col {idx+1} ({openpyxl.utils.get_column_letter(idx+1)}): {col_name}")
            break
            
    if header_row_index:
        print(f"\nData likely starts at Row {header_row_index + 1}")
    else:
        print("\nCould not automatically detect standard header row.")

except Exception as e:
    print(f"Error analyzing file: {e}")
